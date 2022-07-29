import os
import telebot
import logging
import openpyxl
from config import *
from flask import Flask, request
from telebot import types

bot = telebot.TeleBot(BOT_TOKEN)
server = Flask(__name__)
logger = telebot.logger
logger.setLevel(logging.DEBUG)

path = 'users.xlsx'
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
user_id = ""
username = ""
A = 0
name = ""
age = 0
height = 0
weight = 0
sex = ""
allergy = ""


@bot.message_handler(commands=['start'])
def start(message):
    global username
    username = message.from_user.username
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    yes = types.KeyboardButton("–•–æ—Ä–æ—à–æ")
    markup.add(yes)
    bot.reply_to(message,
                 f'–î–æ–±—Ä–æ –ø–æ–∂–∞–ª–æ–≤–∞—Ç—å, {username}. –ü–µ—Ä–µ–¥ —Ä–∞–±–æ—Ç–æ–π —Å –±–æ—Ç–æ–º, –í—ã –¥–æ–ª–∂–Ω—ã –æ—Ç–≤–µ—Ç–∏—Ç—å –Ω–∞ –Ω–µ—Å–∫–æ–ª—å–∫–æ –≤–æ–ø—Ä–æ—Å–æ–≤.',
                 reply_markup=markup)


@bot.message_handler(content_types=['text'])
def reg(message):
    markup_close = types.ReplyKeyboardRemove()
    bot.send_message(message.chat.id, "–ü—Ä–µ–¥–ª–∞–≥–∞—é –ø–æ–∑–Ω–∞–∫–æ–º–∏—Ç—å—Å—è! –ö–∞–∫ –≤–∞—Å –∑–æ–≤—É—Ç?", reply_markup=markup_close)
    bot.register_next_step_handler(message, reg_name)


def reg_name(message):
    global name
    name = message.text
    bot.reply_to(message, "–ü—Ä–µ–∫—Ä–∞—Å–Ω–æ–µ –∏–º—è!")
    bot.send_message(message.chat.id, "–°–∫–æ–ª—å–∫–æ –í–∞–º –ª–µ—Ç?")
    bot.register_next_step_handler(message, reg_age)


def reg_age(message):
    global age
    age = int(message.text)
    bot.reply_to(message, "–û–∫–µ–π")
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    button1 = types.KeyboardButton("–ñ–µ–Ω—Å–∫–∏–π")
    button2 = types.KeyboardButton("–ú—É–∂—Å–∫–æ–π")
    markup.add(button1, button2)
    bot.send_message(message.chat.id, "–í–∞—à –ø–æ–ª:", reply_markup=markup)
    bot.register_next_step_handler(message, reg_sex)


def reg_sex(message):
    global sex
    sex = message.text
    global f
    if sex == "–ñ–µ–Ω—Å–∫–∏–π":
        f = -161
    else:
        f = 5
    bot.send_message(message.chat.id, "–í–∞—à —Ä–æ—Å—Ç –≤ —Å–∞–Ω—Ç–∏–º–µ—Ç—Ä–∞—Ö?", reply_markup=types.ReplyKeyboardRemove())
    bot.register_next_step_handler(message, reg_height)


def reg_height(message):
    global height
    height = int(message.text)
    bot.send_message(message.chat.id, "–°–∫–æ–ª—å–∫–æ –í—ã –≤–µ—Å–∏—Ç–µ –≤ –∫–∏–ª–ª–æ–≥—Ä–∞–º–º–∞—Ö?")
    bot.register_next_step_handler(message, reg_weight)
    

def reg_weight(message):
    global weight
    weight = int(message.text)
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    button1 = types.KeyboardButton("–ú–∏–Ω–∏–º–∞–ª—å–Ω–∞—è –∞–∫—Ç–∏–≤–Ω–æ—Å—Ç—å")
    button2 = types.KeyboardButton("–°–ª–∞–±–∞—è –∞–∫—Ç–∏–≤–Ω–æ—Å—Ç—å: —Ä–∞–∑ –≤ –Ω–µ–¥–µ–ª—é")
    button3 = types.KeyboardButton("–°—Ä–µ–¥–Ω—è—è –∞–∫—Ç–∏–≤–Ω–æ—Å—Ç—å: 3 —Ä–∞–∑–∞ –≤ –Ω–µ–¥–µ–ª—é")
    button4 = types.KeyboardButton("–í—ã—Å–æ–∫–∞—è –∞–∫—Ç–∏–≤–Ω–æ—Å—Ç—å: –ø–æ—á—Ç–∏ –∫–∞–∂–¥—ã–π –¥–µ–Ω—å")
    button5 = types.KeyboardButton("–≠–∫—Å—Ç—Ä–∞-–∞–∫—Ç–∏–≤–Ω–æ—Å—Ç—å: —Ç—è–∂–µ–ª–∞—è —Ñ–∏–∑–∏—á–µ—Å–∫–∞—è —Ä–∞–±–æ—Ç–∞; —Å–ø–æ—Ä—Ç")
    markup.add(button1, button2, button3, button4, button5)
    bot.send_message(message.chat.id, "–í–∞—à–∞ —Å—Ç–µ–ø–µ–Ω—å —Ñ–∏–∑–∏—á–µ—Å–∫–æ–π –∞–∫—Ç–∏–≤–Ω–æ—Å—Ç–∏:", reply_markup=markup)
    bot.register_next_step_handler(message, reg_all)
    
    
def reg_all(message):
    markup_inline = types.InlineKeyboardMarkup(row_width=3)
    item1 = types.InlineKeyboardButton(text='‚ûñ–ú–æ–ª–æ–∫–æ', callback_data='–º–æ–ª–æ–∫–æ')
    item2 = types.InlineKeyboardButton(text='‚ûñ–Ø–π—Ü–æ', callback_data='—è–π—Ü–æ')
    item3 = types.InlineKeyboardButton(text='‚ûñ–ü—à–µ–Ω–∏—Ü–∞', callback_data='–ø—à–µ–Ω–∏—Ü–∞')
    item4 = types.InlineKeyboardButton(text='‚ûñ–†—ã–±–∞', callback_data='—Ä—ã–±–∞')
    item5 = types.InlineKeyboardButton(text='‚ûñ–û—Ä–µ—Ö–∏', callback_data='–æ—Ä–µ—Ö–∏')
    item6 = types.InlineKeyboardButton(text='‚ûñ–ì—Ä–∏–±—ã', callback_data='–≥—Ä–∏–±—ã')
    item7 = types.InlineKeyboardButton(text="‚ûñ–ö—É—Ä–∏—Ü–∞", callback_data='–∫—É—Ä–∏—Ü–∞')
    item8 = types.InlineKeyboardButton(text='‚ûñ–®–æ–∫–æ–ª–∞–¥', callback_data='—à–æ–∫–æ–ª–∞–¥')
    item9 = types.InlineKeyboardButton(text='‚ûñ–ö–æ—Ñ–µ', callback_data='–∫–æ—Ñ–µ')
    item10 = types.InlineKeyboardButton(text='‚ûñ–ö–∞—Ä—Ç–æ—Ñ–µ–ª—å', callback_data='–∫–∞—Ä—Ç–æ—Ñ–µ–ª—å')
    item11 = types.InlineKeyboardButton(text='‚ûñ–õ–∏–º–æ–Ω', callback_data='–ª–∏–º–æ–Ω')
    item12 = types.InlineKeyboardButton(text='‚ûñ–†–∏—Å', callback_data='—Ä–∏—Å')
    item13 = types.InlineKeyboardButton(text='‚ûñ–ü–µ—Ä–µ—Ü', callback_data='–ø–µ—Ä–µ—Ü')
    item14 = types.InlineKeyboardButton(text='–û–∫', callback_data='–æ–∫')
    markup_inline.add(item1, item2, item3, item4, item5, item6, item7, item8, item9, item10, item11, item12, item13, item14)
    bot.send_message(message.chat.id, "–ù–∞ –∫–∞–∫–∏–µ –ø—Ä–æ–¥—É–∫—Ç—ã —É –í–∞—Å –∞–ª–ª–µ—Ä–≥–∏—è?", reply_markup=markup_inline)
    global allergy


@bot.callback_query_handler(func=lambda call: True)
def callback(call):
     products = ["–º–æ–ª–æ–∫–æ", "—è–π—Ü–æ", "–ø—à–µ–Ω–∏—Ü–∞", "—Ä—ã–±–∞", "–æ—Ä–µ—Ö–∏", "–≥—Ä–∏–±—ã", "–∫—É—Ä–∏—Ü–∞", "—à–æ–∫–æ–ª–∞–¥", "–∫–æ—Ñ–µ", "–∫–∞—Ä—Ç–æ—Ñ–µ–ª—å",
                "–ª–∏–º–æ–Ω", "—Ä–∏—Å", "–ø–µ—Ä–µ—Ü"]
     for i in range(0, 14):
        if call.data == products[i]:
            bot.send_message(call.message.chat.id, str(i) +"." + call.data)
            allergy = allergy + call.data + ","
     if call.data == "–æ–∫":
        bot.send_message(call.message.chat.id, '–•–æ—Ä–æ—à–æ!')
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
        button1 = types.KeyboardButton("–ü–æ–¥—Ç–≤–µ—Ä–∂–¥–∞—é")
        markup.add(button1)
        bot.send_message(call.message.chat.id, "–í—ã —É–≤–µ—Ä–µ–Ω—ã?", reply_markup=markup)


@bot.message_handler(content_types=['text'])
def reg_phy(message):
    global phy
    global A
    phy = message.text
    bot.reply_to(message, 'C–ø–∞—Å–∏–±–æ –∑–∞ –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—é!', reply_markup=types.ReplyKeyboardRemove())
    if phy == "–ú–∏–Ω–∏–º–∞–ª—å–Ω–∞—è –∞–∫—Ç–∏–≤–Ω–æ—Å—Ç—å":
        A = 1.2
    elif phy == "–°–ª–∞–±–∞—è –∞–∫—Ç–∏–≤–Ω–æ—Å—Ç—å: —Ä–∞–∑ –≤ –Ω–µ–¥–µ–ª—é":
        A = 1.375
    elif phy == "–°—Ä–µ–¥–Ω—è—è –∞–∫—Ç–∏–≤–Ω–æ—Å—Ç—å: 3 —Ä–∞–∑–∞ –≤ –Ω–µ–¥–µ–ª—é":
        A = 1.55
    elif phy == "–í—ã—Å–æ–∫–∞—è –∞–∫—Ç–∏–≤–Ω–æ—Å—Ç—å: –ø–æ—á—Ç–∏ –∫–∞–∂–¥—ã–π –¥–µ–Ω—å":
        A = 1.725
    elif phy == "–≠–∫—Å—Ç—Ä–∞-–∞–∫—Ç–∏–≤–Ω–æ—Å—Ç—å: —Ç—è–∂–µ–ª–∞—è —Ñ–∏–∑–∏—á–µ—Å–∫–∞—è —Ä–∞–±–æ—Ç–∞; —Å–ø–æ—Ä—Ç":
        A = 1.9
    else:
        bot.send_message(message.chat.id, "invalid data")
    global call
    call = (10 * weight + 6.25 * height - 5 * age + f) * A
    bot.send_message(message.chat.id,
                     "–ë–æ—Ç —Ä–∞—Å—á–∏—Ç—ã–≤–∞–µ—Ç –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ –∫–∞–ª–æ—Ä–∏–π –ø–æ —Ñ–æ—Ä–º—É–ª–µ –ú–∏—Ñ—Ñ–ª–∏–Ω–∞-–°–∞–Ω –ñ–µ–æ—Ä–∞- –æ–¥–Ω–æ–π –∏–∑ —Å–∞–º—ã—Ö –ø–æ—Å–ª–µ–¥–Ω–∏—Ö —Ñ–æ—Ä–º—É–ª —Ä–∞—Å—á–µ—Ç–∞ –∫–∞–ª–æ—Ä–∏–π –¥–ª—è –æ–ø—Ç–∏–º–∞–ª—å–Ω–æ–≥–æ –ø–æ—Ö—É–¥–µ–Ω–∏—è –∏–ª–∏ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏—è –Ω–æ—Ä–º–∞–ª—å–Ω–æ–≥–æ –≤–µ—Å–∞.")
    bot.send_message(message.chat.id,
                     "–ù–µ–æ–±—Ö–æ–¥–∏–º–æ–µ –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ –∫–∏–ª–æ–∫–∞–ª–æ—Ä–∏–π (–∫–∫–∞–ª) –≤ —Å—É—Ç–∫–∏ –¥–ª—è –í–∞—Å = " + str(call) + " " + "–∫–∫–∞–ª")
            

@bot.message_handler(content_types=['text'])
def reg_save(message): 
    bot.send_message(message.chat.id, "–ù–∞–∂–º–∏—Ç–µ /save –¥–ª—è —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏—è –¥–∞–Ω–Ω—ã—Ö ")               
                
                
@bot.message_handler(commands=['save'])
def save(message):
    bot.send_message(message.chat.id, '–î–∞–Ω–Ω—ã–µ —Å–æ—Ö—Ä–∞–Ω–µ–Ω—ã')
    sheet_obj.cell(row=sheet_obj.max_row + 1, column=1).value = user_id
    sheet_obj.cell(row=sheet_obj.max_row + 1, column=2).value = username
    sheet_obj.cell(row=sheet_obj.max_row + 1, column=3).value = name
    sheet_obj.cell(row=sheet_obj.max_row + 1, column=4).value = age
    sheet_obj.cell(row=sheet_obj.max_row + 1, column=5).value = height
    sheet_obj.cell(row=sheet_obj.max_row + 1, column=6).value = weight
    sheet_obj.cell(row=sheet_obj.max_row + 1, column=7).value = phy
    sheet_obj.cell(row=sheet_obj.max_row + 1, column=8).value = call
    sheet_obj.cell(row=sheet_obj.max_row + 1, column=9).value = allergy


@bot.message_handler(commands=['receipt'])
def receipt(message):
    bot.send_message(message.chat.id, '–ó–¥–æ—Ä–æ–≤–æ–µ –ø–∏—Ç–∞–Ω–∏–µ –æ–±–µ—Å–ø–µ—á–∏–≤–∞–µ—Ç –Ω–æ—Ä–º–∞–ª—å–Ω–æ–µ —Ä–∞–∑–≤–∏—Ç–∏–µ –∏ –∂–∏–∑–Ω–µ–¥–µ—è—Ç–µ–ª—å–Ω–æ—Å—Ç—å —á–µ–ª–æ–≤–µ–∫–∞, —Å–ø–æ—Å–æ–±—Å—Ç–≤—É–µ—Ç —É–∫—Ä–µ–ø–ª–µ–Ω–∏—é –µ–≥–æ –∑–¥–æ—Ä–æ–≤—å—è –∏ –ø—Ä–æ—Ñ–∏–ª–∞–∫—Ç–∏–∫–µ –∑–∞–±–æ–ª–µ–≤–∞–Ω–∏–π. –ó–¥–æ—Ä–æ–≤–æ–µ –ø–∏—Ç–∞–Ω–∏–µ ‚àí –∑–∞–ª–æ–≥ –¥–æ–ª–≥–æ–π –∂–∏–∑–Ω–∏. –î–ª—è —Ö–æ—Ä–æ—à–µ–≥–æ —Å–∞–º–æ—á—É–≤—Å—Ç–≤–∏—è, –ø—Ä–∞–≤–∏–ª—å–Ω–æ–≥–æ —Ñ—É–Ω–∫—Ü–∏–æ–Ω–∏—Ä–æ–≤–∞–Ω–∏—è –æ—Ä–≥–∞–Ω–æ–≤ –Ω–∞–º –Ω–µ–æ–±—Ö–æ–¥–∏–º–æ –ø–∏—Ç–∞—Ç—å—Å—è –∫–∞—á–µ—Å—Ç–≤–µ–Ω–Ω–æ–π –ø–∏—â–µ–π. –†–∞–∑–Ω–æ–æ–±—Ä–∞–∑–∏–µ —Ñ—Ä—É–∫—Ç–æ–≤ –∏ –æ–≤–æ—â–µ–π –Ω–∞ –≤–∞—à–µ–º —Å—Ç–æ–ª–µ, –∫–∏—Å–ª–æ–º–æ–ª–æ—á–Ω–∞—è –∏ –º–æ–ª–æ—á–Ω–∞—è –ø—Ä–æ–¥—É–∫—Ü–∏—è, –æ—Ä–µ—Ö–∏ –∏ –∑–ª–∞–∫–∏, —ç—Ç–æ –∏ –µ—Å—Ç—å –∑–¥–æ—Ä–æ–≤–æ–µ –ø–∏—Ç–∞–Ω–∏–µ! –ú—ã –ø–æ—Å—Ç–∞—Ä–∞–ª–∏—Å—å —Å–æ–±—Ä–∞—Ç—å –¥–ª—è –≤–∞—Å –ª—É—á—à–∏–µ —Ä–µ—Ü–µ–ø—Ç—ã —Å–∞–ª–∞—Ç–æ–≤, –ø–µ—Ä–≤—ã—Ö –∏ –≤—Ç–æ—Ä—ã—Ö –±–ª—é–¥, –¥–µ—Å–µ—Ä—Ç–æ–≤ –∏ –≤—ã–ø–µ—á–∫–∏ —Å –º–∞–∫—Å–∏–º–∞–ª—å–Ω–æ–π –ø–æ–ª—å–∑–æ–π –¥–ª—è –≤–∞—à–µ–≥–æ –∑–¥–æ—Ä–æ–≤—å—è! üíö')
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    button1 = types.KeyboardButton(text='–†–µ—Ü–µ–ø—Ç—ã —Å–∞–ª–∞—Ç–æ–≤')
    button2 = types.KeyboardButton(text='–ü–µ—Ä–≤—ã–µ –±–ª—é–¥–∞')
    button3 = types.KeyboardButton(text='–í—Ç–æ—Ä—ã–µ –±–ª—é–¥–∞')
    button4 = types.KeyboardButton(text='–ù–∞—Ü–∏–æ–Ω–∞–ª—å–Ω—ã–µ –±–ª—é–¥–∞')
    button5 = types.KeyboardButton(text='–í—ã–ø–µ—á–∫–∞')
    button6 = types.KeyboardButton(text='–†–∞–∑–Ω–æ–µ')
    button7 = types.KeyboardButton(text='–î–µ—Å–µ—Ä—Ç—ã –∏ —Å–ª–∞–¥–æ—Å—Ç–∏')
    button8 = types.KeyboardButton(text='–ó–∞–≥–æ—Ç–æ–≤–∫–∏, —Å–æ–ª–µ–Ω—å—è, –≤–∞—Ä–µ–Ω—å—è')
    button9 = types.KeyboardButton(text='–°–æ—É—Å—ã')
    button10 = types.KeyboardButton(text='–†–µ—Ü–µ–ø—Ç—ã –¥–ª—è –º—É–ª—å—Ç–∏–≤–∞—Ä–æ–∫')
    markup.row(button1, button2, button3, button4, button5)
    markup.row(button6, button7, button8, button9, button10)


@bot.message_handler(content_types=['text'])
def user_text(message):
    if message.text.lower() == '–º–æ—è –ø–µ—Ä—Å–æ–Ω–∞–ª—å–Ω–∞—è –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—è':
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
        button1 = types.KeyboardButton(text='–ò–º—è')
        button2 = types.KeyboardButton(text='–í–æ–∑—Ä–∞—Å—Ç')
        button3 = types.KeyboardButton(text='–ü–æ–ª')
        button4 = types.KeyboardButton(text='–†–æ—Å—Ç')
        button5 = types.KeyboardButton(text='–í–µ—Å')
        button6 = types.KeyboardButton(text='–§–∏–∑–∏—á–µ—Å–∫–∞—è –∞–∫—Ç–∏–≤–Ω–æ—Å—Ç—å')
        button7 = types.KeyboardButton(text='–ö–∫–∞–ª –≤ —Å—É—Ç–∫–∏')
        button8 = types.KeyboardButton(text='–ù–∞–∑–∞–¥')
        markup.add(button1, button2, button3, button4)
        markup.add(button5, button6, button7, button8)
        bot.send_message(message.chat.id, '', reply_markup=markup)


@bot.message_handler(content_types=['text'])
def get_user_book(message):
    book = message.text.lower()
    path = open("/content/drive/MyDrive/NIS —É—á–µ–±–Ω–∏–∫–∏/" + book + ".pdf", "rb")
    bot.send_message(message.chat.id, "–ò—â–µ–º –∫–Ω–∏–≥—É...")
    bot.send_document(message.chat.id, path)
    path.close()


@server.route(f"/{BOT_TOKEN}", methods=["POST"])
def redirect_message():
    json_string = request.get_data().decode("utf-8")
    update = telebot.types.Update.de_json(json_string)
    bot.process_new_updates([update])
    return "!", 200


if __name__ == "__main__":
    bot.remove_webhook()
    bot.set_webhook(url=APP_URL)
    server.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)))


